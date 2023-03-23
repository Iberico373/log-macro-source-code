import os
import json
import openpyxl
import xlwings as xw
import io
import msoffcrypto

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def Init():
    #Decrypt workbook password in order to access it
    decryptedWorkbook = io.BytesIO()

    with open('insertExcelSheetNameHere', 'rb') as file:
        officeFile = msoffcrypto.OfficeFile(file)
        officeFile.load_key(password="insertPassHere")
        officeFile.decrypt(decryptedWorkbook)

    #Init variables
    global workbook 
    global prepostSheet 
    global interactionSheet 
    global ptpDataEntries
    global ptpDataListPrePost
    global ptpDataListInteraction 
    global previousDateArr

    workbook = openpyxl.load_workbook(filename=decryptedWorkbook)
    prepostSheet = workbook["Pre-post"]
    interactionSheet = workbook["Interaction"]
    ptpDataEntries = []
    ptpDataListPrePost = []
    ptpDataListInteraction = []
    previousDateArr = []

#Gets files from current directory and stores them in ptpDataList
def GetFiles():
    for file in os.listdir():
        if ".json" in file:
            os.rename(file, str(file.lower()))
            ptpDataEntries.append(file)
            ptpDataEntries.sort()

def StripLeadingZero(day, month, year):   
    day = day.lstrip('0')
    month = month.lstrip('0')
    year = year.lstrip('0')

    date = day + "/" + month + "/" + year

    return date

def FormatDate(startDate, currentDate = ""):
    startDate = StripLeadingZero(startDate.split("/")[0], startDate.split("/")[1], startDate.split("/")[2])
    startDateArr = startDate.split("/")

    day = ""
    month = ""
    year = ""
    highlight = False

    if (currentDate != ""):
        currentDate = StripLeadingZero(currentDate.split("/")[0], currentDate.split("/")[1], currentDate.split("/")[2])
        currentDateArr = currentDate.split("/")

        for i in range(3):
            #print(currentDateArr[i] + " - " + startDateArr[i] + " = " + str(int(currentDateArr[i]) - int(startDateArr[i])))
            if (startDateArr[i] != currentDateArr[i]):
                if(abs(int(currentDateArr[i]) - int(startDateArr[i])) > 2):        
                    if (month == ""):
                        month = day

                    day = startDateArr[i]   

                else:
                    if (day == ""):
                        day = startDateArr[i]

                    elif (month == ""):
                        month = startDateArr[i]
            
            else:
                if (int(startDateArr[i]) > 12):
                    year = startDateArr[i]

                else:
                    month = startDateArr[i]
    

    else:
        for i in range(3):
            if (int(startDateArr[i]) > 1000):
                year = startDateArr[i]

            elif (int(startDateArr[i]) > 3):
                if (month == ""):
                    month = day

                day = startDateArr[i]    
            
            else:
                if (day == ""):
                    day = startDateArr[i]

                month = startDateArr[i]

    date = day + "/" + month + "/" + year

    if (int(day) < 4 and int(day) != int(month)):
        highlight = True

    dateList = [date, highlight]

    return dateList

#Parse data from json to be appended into excel        
def ParseData():
    for file in ptpDataEntries:
        #Instantiate two dictionaries for 'Pre-post' sheet
        #and 'Interaction' sheet
        ptpDataDictPrePost = {
        'participantId' : "",
        'preQ1' : "",
        'preQ2' : "",
        'preQ3' : "",
        'preQ4' : "",
        'preQ5' : "",
        'preQ6' : "",
        'preQ7' : "",
        'preQ8' : "",
        'preQ9' : "",
        'preQ10' : "",
        'preQ11' : "",
        'postQ1' : "",
        'postQ2' : "",
        'postQ3' : "",
        'postQ4' : "",
        'postQ5' : "",
        'postQ6' : "",
        'postQ7' : "",
        'postQ8' : "",
        'postQ9' : "",
        'postQ10' : "",
        'postQ11' : "",
        'parkLittleBoy' : 0, 
        'parkBillBoard' : 0,
        'supermarketLittleGirl' : 0,
        'supermarketBillboard' : 0,
        'trainStationGirl' : 0,
        'trainStationBulletinBoard' : 0
        }

        date = {
            'startDate' : "",
            'highlight' : False
        }

        ptpDataDictInteraction = {
            'date' : date,
            'participantId' : "",
            'timeSpent' : 0
        }

        participantId = file.removesuffix("_activities.json").upper()
        ptpDataDictPrePost['participantId'] = participantId
        ptpDataDictInteraction['participantId'] = participantId
        startDate = ""

        data = json.load(open(file))

        for activities in data:
            if activities['activity']['name'] == "GAME_TIME":
                startDate = (str(activities['activity']['startTime']).split(" ")[0])
                ptpDataDictInteraction['timeSpent'] = activities['activity']['totalTimeSpent']

            elif activities['activity']['name'] == "FIRST_LOGIN":
                startDate = (str(activities['activity']['startTime']).split(" ")[0])

            elif activities['activity']['name'] == "DAILY_LOGIN" and ptpDataDictInteraction['date']['startDate'] == "":
                currentDate = (str(activities['activity']['startTime']).split(" ")[0])

                if (startDate != currentDate):
                    dateList = FormatDate(startDate, currentDate)
                    date['startDate'] = dateList[0]
                    date['highlight'] = dateList[1]

                    ptpDataDictInteraction['date'] = date
            
            elif activities['activity']['name'] == "PREQUIZ":
                for response in activities['responses']:
                    currentKey = list(ptpDataDictPrePost)[int(response['responseIndex']) + 1]

                    if bool(response['responseType']):
                        ptpDataDictPrePost[currentKey] = "Wrong"
                    
                    else:
                        ptpDataDictPrePost[currentKey] = "Correct"

            elif activities['activity']['name'] == "POSTQUIZ":
                for response in activities['responses']:
                    currentKey = list(ptpDataDictPrePost)[int(response['responseIndex']) + 12]

                    if bool(response['responseType']):
                        ptpDataDictPrePost[currentKey] = "Wrong"
                    
                    else:
                        ptpDataDictPrePost[currentKey] = "Correct"         

            elif activities['activity']['name'] == "NPC_INTERACTION":
                if activities['gameDetailId'] == "29c9c745-6cf9-44a4-a0d3-b860d2091382":
                    ptpDataDictPrePost['parkLittleBoy'] = 1
                
                elif activities['gameDetailId'] == "4bf0895c-fd45-4fa6-bb17-6652c2e0bafa":
                    ptpDataDictPrePost['parkBillBoard'] = 1
                
                elif activities['gameDetailId'] == "da6df25e-90eb-43e9-9bd7-007b16102075":
                    ptpDataDictPrePost['supermarketLittleGirl'] = 1
                
                elif activities['gameDetailId'] == "08732bf5-5467-48cd-b9d3-ff0df2de0994":
                    ptpDataDictPrePost['supermarketBillboard'] = 1
                
                elif activities['gameDetailId'] == "b963c2fa-7dd5-40e5-9477-7acf38a9de73":
                    ptpDataDictPrePost['trainStationGirl'] = 1
                
                elif activities['gameDetailId'] == "7c3c868a-336e-4ef1-bd49-e85fdf3489ea":
                    ptpDataDictPrePost['trainStationBulletinBoard'] = 1

        if (ptpDataDictInteraction['date']['startDate'] == ""):
            dateList = FormatDate(startDate)
            date['startDate'] = dateList[0]
            date['highlight'] = dateList[1]
            
            ptpDataDictInteraction['date'] = date

        ptpDataListPrePost.append(ptpDataDictPrePost)
        ptpDataListInteraction.append(ptpDataDictInteraction)


def AppendDataToTable():
    rowIndex = 2
    colIndex = 1

    for data in ptpDataListPrePost:
        for value in data.values():
            currentCell = str(get_column_letter(colIndex)) + str(rowIndex) 
            prepostSheet[currentCell].fill = PatternFill(fill_type='none')
            prepostSheet[currentCell] = value
            colIndex = colIndex + 1

        rowIndex = rowIndex + 1
        colIndex = 1

    rowIndex = 2
    colIndex = 1

    for data in ptpDataListInteraction:
        for value in data.values():
            currentCell = str(get_column_letter(colIndex)) + str(rowIndex)

            if (type(value) is dict):
                interactionSheet[currentCell] = value['startDate']

                if (value['highlight']):
                    for i in range(3):
                        cell = str(get_column_letter(colIndex + i)) + str(rowIndex) 
                        interactionSheet[cell].fill = PatternFill(start_color='FFFF00', end_color='FFFF00',fill_type='solid')

            else:
                interactionSheet[currentCell] = value
                #interactionSheet[currentCell].fill = PatternFill(fill_type='none')
            
            colIndex = colIndex + 1

        rowIndex = rowIndex + 1
        colIndex = 1
    
def SaveWorkbook():
    workbook.save('Catalyst app data.xlsx')
    workbook.close()
    wb = xw.Book('Catalyst app data.xlsx')
    wb.save(password = 'insertPassHere', path = 'insertPathToExcelSheetHere')

def main():
    Init()
    GetFiles()
    ParseData()  
    AppendDataToTable()
    print('Total participants logged: ' + str(len(ptpDataEntries)))
    SaveWorkbook()

if __name__ == "__main__":
    main()